import io
from datetime import datetime

from django.http import HttpResponse
from django.views import View

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate,
    Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.enums import TA_LEFT, TA_RIGHT

# Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.expenses.expenses_model import Expenses, ExpensesItems

# ── Company constants ─────────────────────────────────────────────────
COMPANY_NAME  = "Arrolite"
COMPANY_FULL  = "Arrolite Group of Companies"
COMPANY_EMAIL = "accounts@arrolite.com"

# ── Colour palette ────────────────────────────────────────────────────
WHITE     = colors.white
NAVY      = colors.HexColor("#0D1B2A")
ACCENT    = colors.HexColor("#1565C0")
ACCENT2   = colors.HexColor("#FFA000")
ALT_ROW   = colors.HexColor("#EEF2F7")
LIGHT_BG  = colors.HexColor("#F5F7FA")
RULE      = colors.HexColor("#CFD8DC")
MID_GREY  = colors.HexColor("#546E7A")
RED_TEXT  = colors.HexColor("#B71C1C")
GRN_TEXT  = colors.HexColor("#1B5E20")


# ── Helpers ───────────────────────────────────────────────────────────
def _parse_date(s):
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date: {s!r}")


def _sty(name, **kw):
    return ParagraphStyle(name, **kw)


def _status_style(status):
    colour_map = {
        "PAID":      colors.HexColor("#1B5E20"),
        "PARTIAL":   colors.HexColor("#E65100"),
        "PENDING":   colors.HexColor("#B71C1C"),
        "OVERDUE":   colors.HexColor("#880E4F"),
        "EXPIRED":   colors.HexColor("#4A148C"),
        "CANCELLED": colors.HexColor("#37474F"),
    }
    c = colour_map.get(status.upper(), MID_GREY)
    return _sty(
        f"ST_{status}",
        fontName="Helvetica-Bold",
        fontSize=7,
        textColor=c,
        leading=10,
    )


# ══════════════════════════════════════════════════════════════════════
class ExpenseExportView(View):

    # ── POST entry-point ──────────────────────────────────────────────
    def post(self, request):
        export_type   = request.POST.get("export_type",   "pdf")
        expenses_type = request.POST.get("expenses_type", "").strip()
        due_date_str  = request.POST.get("due_date",      "").strip() 
        filters = {}
        if expenses_type:
            filters["expenses__invoice_number"] = expenses_type

        if due_date_str:
            if "to" in due_date_str:
                start_str, end_str = [d.strip() for d in due_date_str.split("to", 1)]
                try:
                    filters["due_date__range"] = (
                        _parse_date(start_str),
                        _parse_date(end_str),
                    )
                except ValueError:
                    pass
            else:
                try:
                    filters["due_date"] = _parse_date(due_date_str)
                except ValueError:
                    pass

        qs = (
            ExpensesItems.objects
            .select_related("expenses", "expenses__expenses_type")
            .filter(**filters)
            .order_by("-created_at")
        )

        rows = []
        for exp in qs:
            expense = exp.expenses
            rows.append({
                "company_name":   expense.company_name   or "—",
                "product_name":   expense.product_name   or "—",
                "invoice_number": expense.invoice_number or "—",
                "invoice_date":   str(expense.due_date)  if expense.due_date  else "—",
                "paid_date":      str(exp.due_date)      if exp.due_date      else "—",
                "expense_status": (expense.expense_status or "PENDING").upper(),
                "expense_type":   expense.expenses_type.name if expense.expenses_type else "—",
                "payment_mode":   (exp.payment_mode or "—").upper(),
                "amount":         float(exp.amount           or 0),   # item paid amount
                "total_amount":   float(expense.amount       or 0),   # expense total
                "total_paid":     float(expense.total_paid() or 0),
                "balance":        float(expense.balance_amount() or 0),
            })

        if export_type == "excel":
            return self._excel(rows)
        return self._pdf(rows)

    # ════════════════════════════════════════════════════════════════
    #  PDF EXPORT
    # ════════════════════════════════════════════════════════════════
    def _pdf(self, rows):
        buf = io.BytesIO()

        doc = BaseDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=0.7*cm, rightMargin=0.7*cm,
            topMargin=2.7*cm,  bottomMargin=1.3*cm,
        )
        frame = Frame(
            doc.leftMargin, doc.bottomMargin,
            doc.width, doc.height, id="main",
        )
        doc.addPageTemplates([
            PageTemplate(
                id="base", frames=frame,
                onPage=lambda c, d: (self._pdf_header(c, d), self._pdf_footer(c, d)),
            )
        ])

        # ── Paragraph styles ──────────────────────────────────────
        H_STYLE = _sty("H",  fontName="Helvetica-Bold", fontSize=7.5,
                        textColor=WHITE, leading=10, alignment=TA_LEFT)
        C_STYLE = _sty("C",  fontName="Helvetica", fontSize=7.5,
                        textColor=colors.HexColor("#1A1A2E"), leading=10)
        TOT_L   = _sty("TL", fontName="Helvetica-Bold", fontSize=8,
                        textColor=WHITE, alignment=TA_LEFT)
        TOT_R   = _sty("TR", fontName="Helvetica-Bold", fontSize=8,
                        textColor=WHITE, alignment=TA_RIGHT)

        # ── Column widths (must sum to doc.width) ─────────────────
        # Col:  1     2     3     4     5     6     7     8     9    10    11    12
        #       SNo  Comp  Prod  InvNo InvDt InvAmt PdDt  Mode PdAmt Stat  Paid  Bal
        W = doc.width
        col_w = [
            W * 0.04,   # 1  S.No
            W * 0.11,   # 2  Company Name
            W * 0.10,   # 3  Product
            W * 0.09,   # 4  Invoice Number
            W * 0.08,   # 5  Invoice Date
            W * 0.09,   # 6  Invoice Amount
            W * 0.08,   # 7  Paid Date
            W * 0.08,   # 8  Payment Mode
            W * 0.08,   # 9  Paid Amount
            W * 0.07,   # 10 Status
            W * 0.08,   # 11 Total Paid
            W * 0.10,   # 12 Balance  ← absorbs rounding
        ]
        col_w[-1] = W - sum(col_w[:-1])   # ✅ guarantees exact fit

        # ── 12 headers — same order as columns ────────────────────
        HDR = [
            "S.No",            # 1
            "Company Name",    # 2
            "Product",         # 3
            "Invoice Number",     # 4
            "Invoice Date",    # 5
            "Invoice Amount",     # 6
            "Paid Date",       # 7
            "Payment Mode",    # 8
            "Paid Amt",        # 9
            "Status",          # 10
            "Total Paid",      # 11
            "Balance",         # 12
        ]
        table_data = [[Paragraph(h, H_STYLE) for h in HDR]]
        row_styles = []

        total_item_amt = total_exp_amt = total_paid = total_bal = 0
        ri = 1   # row index in table (0 = header)

        for ei, row in enumerate(rows):
            bg = ALT_ROW if ei % 2 == 0 else WHITE

            amt_sty = _sty(f"A{ei}", fontName="Helvetica-Bold", fontSize=7.5,
                           textColor=NAVY,     alignment=TA_RIGHT, leading=10)
            pay_sty = _sty(f"P{ei}", fontName="Helvetica",      fontSize=7.5,
                           textColor=MID_GREY, alignment=TA_RIGHT, leading=10)
            bal_sty = _sty(f"B{ei}", fontName="Helvetica-Bold", fontSize=7.5,
                           textColor=RED_TEXT if row["balance"] > 0 else GRN_TEXT,
                           alignment=TA_RIGHT, leading=10)

            # ✅ 12 cells — exact same order as HDR
            table_data.append([
                Paragraph(str(ei + 1),                              C_STYLE),   # 1
                Paragraph(f"<b>{row['company_name']}</b>",          C_STYLE),   # 2
                Paragraph(row["product_name"],                      C_STYLE),   # 3
                Paragraph(row["invoice_number"],                    C_STYLE),   # 4
                Paragraph(row["invoice_date"],                      C_STYLE),   # 5
                Paragraph(f"Rs.{row['total_amount']:,.2f}",         amt_sty),   # 6
                Paragraph(row["paid_date"],                         C_STYLE),   # 7
                Paragraph(row["payment_mode"],                      C_STYLE),   # 8
                Paragraph(f"Rs.{row['amount']:,.2f}",               amt_sty),   # 9
                Paragraph(row["expense_status"],
                          _status_style(row["expense_status"])),                # 10
                Paragraph(f"Rs.{row['total_paid']:,.2f}",           pay_sty),   # 11
                Paragraph(f"Rs.{row['balance']:,.2f}",              bal_sty),   # 12
            ])
            row_styles += [
                ("BACKGROUND", (0, ri), (-1, ri), bg),
                ("LINEBELOW",  (0, ri), (-1, ri), 0.4, RULE),
            ]
            ri += 1
            total_item_amt += row["amount"]
            total_exp_amt  += row["total_amount"]
            total_paid     += row["total_paid"]
            total_bal      += row["balance"]

        # ── Grand-total footer row ────────────────────────────────
        table_data.append([
            Paragraph("",                              TOT_L),   # 1
            Paragraph("GRAND TOTAL",                   TOT_L),   # 2
            Paragraph("",                              TOT_L),   # 3
            Paragraph("",                              TOT_L),   # 4
            Paragraph(f"{len(rows)} records",          TOT_L),   # 5
            Paragraph("",                              TOT_L),   # 6
            Paragraph("",                              TOT_L),   # 7
            Paragraph("",                              TOT_L),   # 8
            Paragraph(f"Rs.{total_item_amt:,.2f}",     TOT_R),   # 9
            Paragraph("",                              TOT_L),   # 10
            Paragraph(f"Rs.{total_paid:,.2f}",         TOT_R),   # 11
            Paragraph(f"Rs.{total_bal:,.2f}",          TOT_R),   # 12
        ])
        row_styles.append(("BACKGROUND", (0, ri), (-1, ri), NAVY))

        tbl_style = TableStyle([
            ("BACKGROUND",    (0,  0), (-1,  0), ACCENT),
            ("LINEBELOW",     (0,  0), (-1,  0), 1.5, ACCENT2),
            ("GRID",          (0,  0), (-1, -1), 0.3, RULE),
            ("BOX",           (0,  0), (-1, -1), 1.0, NAVY),
            ("VALIGN",        (0,  0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0,  0), (-1, -1), 5),
            ("BOTTOMPADDING", (0,  0), (-1, -1), 5),
            ("LEFTPADDING",   (0,  0), (-1, -1), 4),
            ("RIGHTPADDING",  (0,  0), (-1, -1), 4),
            ("TOPPADDING",    (0,  0), (-1,  0), 7),
            ("BOTTOMPADDING", (0,  0), (-1,  0), 7),
        ] + row_styles)

        main_tbl = Table(table_data, colWidths=col_w, repeatRows=1)
        main_tbl.setStyle(tbl_style)

        # ── KPI summary cards (5 cards) ───────────────────────────
        collection_rate = (total_paid / total_item_amt * 100) if total_item_amt else 0
        kpi_data = [[
            Paragraph(
                f"<font size='7' color='#607D8B'>Total Records</font><br/>"
                f"<font size='13' color='#0D1B2A'><b>{len(rows)}</b></font>",
                C_STYLE),
            Paragraph(
                f"<font size='7' color='#607D8B'>Invoice Amount</font><br/>"
                f"<font size='11' color='#0D1B2A'><b>Rs.{total_exp_amt:,.2f}</b></font>",
                C_STYLE),
            Paragraph(
                f"<font size='7' color='#607D8B'>Total Paid</font><br/>"
                f"<font size='11' color='#1B5E20'><b>Rs.{total_paid:,.2f}</b></font>",
                C_STYLE),
            Paragraph(
                f"<font size='7' color='#607D8B'>Outstanding Balance</font><br/>"
                f"<font size='11' color='#B71C1C'><b>Rs.{total_bal:,.2f}</b></font>",
                C_STYLE),
            Paragraph(
                f"<font size='7' color='#607D8B'>Collection Rate</font><br/>"
                f"<font size='11' color='#1565C0'><b>{collection_rate:.1f}%</b></font>",
                C_STYLE),
        ]]
        kpi_tbl = Table(kpi_data, colWidths=[W / 5] * 5)
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), LIGHT_BG),
            ("BOX",           (0, 0), (-1, -1), 0.8, NAVY),
            ("LINEBEFORE",    (1, 0), (-1, -1), 0.5, RULE),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ]))

        doc.build([kpi_tbl, Spacer(0, 0.35*cm), main_tbl])
        buf.seek(0)

        return HttpResponse(
            buf,
            content_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="arrolite_expense_report.pdf"'},
        )

    # ── PDF Header ────────────────────────────────────────────────────
    def _pdf_header(self, canv, doc):
        canv.saveState()
        pw = doc.pagesize[0]
        ph = doc.pagesize[1]

        canv.setFillColor(NAVY)
        canv.rect(0, ph - 2.2*cm, pw, 2.2*cm, stroke=0, fill=1)

        canv.setFillColor(ACCENT2)
        canv.rect(0, ph - 2.2*cm, 0.55*cm, 2.2*cm, stroke=0, fill=1)

        bx, by, bw, bh = 0.85*cm, ph - 1.75*cm, 3.8*cm, 1.2*cm
        canv.setFillColor(WHITE)
        canv.roundRect(bx, by, bw, bh, 4, stroke=0, fill=1)
        canv.setFont("Helvetica-Bold", 11)
        canv.setFillColor(NAVY)
        canv.drawCentredString(bx + bw / 2, by + 0.35*cm, COMPANY_NAME.upper())

        canv.setFont("Helvetica-Bold", 15)
        canv.setFillColor(WHITE)
        canv.drawString(5.3*cm, ph - 1.25*cm, "EXPENSE REPORT")

        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#B0BEC5"))
        canv.drawString(5.3*cm, ph - 1.75*cm, "Financial Summary  ·  All Expense Entries")

        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#90A4AE"))
        canv.drawRightString(pw - 0.6*cm, ph - 1.10*cm,
                             f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
        canv.drawRightString(pw - 0.6*cm, ph - 1.60*cm, f"Page {doc.page}")

        canv.setFillColor(ACCENT2)
        canv.rect(0, ph - 2.2*cm, pw, 0.12*cm, stroke=0, fill=1)
        canv.restoreState()

    # ── PDF Footer ────────────────────────────────────────────────────
    def _pdf_footer(self, canv, doc):
        canv.saveState()
        pw = doc.pagesize[0]

        canv.setFillColor(NAVY)
        canv.rect(0, 0, pw, 0.9*cm, stroke=0, fill=1)

        canv.setFillColor(ACCENT2)
        canv.rect(0, 0, 0.55*cm, 0.9*cm, stroke=0, fill=1)

        canv.setFont("Helvetica", 7)
        canv.setFillColor(colors.HexColor("#90A4AE"))
        canv.drawString(0.85*cm, 0.3*cm,
                        f"System-generated report. Queries: {COMPANY_EMAIL}")
        canv.drawRightString(pw - 0.6*cm, 0.3*cm,
                             f"© {datetime.now().year} {COMPANY_FULL}")
        canv.restoreState()

    # ════════════════════════════════════════════════════════════════
    #  EXCEL EXPORT  ✅ complete
    # ════════════════════════════════════════════════════════════════
    def _excel(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        # ── Styles ────────────────────────────────────────────────
        hdr_fill  = PatternFill("solid", fgColor="0D1B2A")
        alt_fill  = PatternFill("solid", fgColor="EEF2F7")
        tot_fill  = PatternFill("solid", fgColor="1565C0")
        kpi_fill  = PatternFill("solid", fgColor="F5F7FA")

        hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        body_font = Font(name="Calibri", size=9)
        tot_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        kpi_font  = Font(name="Calibri", bold=True, size=11)

        center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        right     = Alignment(horizontal="right",  vertical="center")

        thin = Side(style="thin", color="CFD8DC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        money_fmt = '#,##0.00'
        red_font  = Font(name="Calibri", bold=True, color="B71C1C", size=9)
        grn_font  = Font(name="Calibri", bold=True, color="1B5E20", size=9)

        # ── Title row ─────────────────────────────────────────────
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value         = f"{COMPANY_FULL} — Expense Report"
        title_cell.font          = Font(name="Calibri", bold=True, size=14, color="0D1B2A")
        title_cell.alignment     = center
        title_cell.fill          = PatternFill("solid", fgColor="EEF2F7")
        ws.row_dimensions[1].height = 28

        # ── Generated date ────────────────────────────────────────
        ws.merge_cells("A2:L2")
        date_cell = ws["A2"]
        date_cell.value      = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
        date_cell.font       = Font(name="Calibri", italic=True, size=9, color="546E7A")
        date_cell.alignment  = center
        ws.row_dimensions[2].height = 16

        # ── KPI summary row ───────────────────────────────────────
        total_item_amt = sum(r["amount"]       for r in rows)
        total_exp_amt  = sum(r["total_amount"] for r in rows)
        total_paid     = sum(r["total_paid"]   for r in rows)
        total_bal      = sum(r["balance"]      for r in rows)
        collection_rate = (total_paid / total_item_amt * 100) if total_item_amt else 0

        kpi_labels = [
            ("Total Records",       len(rows),                  None),
            ("Invoice Amount",      total_exp_amt,              money_fmt),
            ("Total Paid",          total_paid,                 money_fmt),
            ("Outstanding Balance", total_bal,                  money_fmt),
            ("Collection Rate",     f"{collection_rate:.1f}%",  None),
        ]

        # KPI spans cols A-L split into 5 groups of ~2 cols each
        kpi_col_spans = ["A3:B4", "C3:D4", "E3:F4", "G3:H4", "I3:L4"]

        for span, (label, value, fmt) in zip(kpi_col_spans, kpi_labels):
            # label row
            label_span = span.replace("4", "3")
            ws.merge_cells(label_span)
            lc = ws[label_span.split(":")[0]]
            lc.value     = label
            lc.font      = Font(name="Calibri", size=8, color="607D8B")
            lc.alignment = center
            lc.fill      = kpi_fill

            # value row
            value_span = span.replace("3", "4")
            ws.merge_cells(value_span)
            vc = ws[value_span.split(":")[0]]
            vc.value     = value
            vc.font      = kpi_font
            vc.alignment = center
            vc.fill      = kpi_fill
            if fmt:
                vc.number_format = fmt

        ws.row_dimensions[3].height = 14
        ws.row_dimensions[4].height = 22

        # ── Header row ────────────────────────────────────────────
        HDR = [
            "S.No", "Company Name", "Product", "Invoice Number",
            "Invoice Date", "Invoice Amount", "Paid Date",
            "Payment Mode", "Paid Amount", "Status", "Total Paid", "Balance",
        ]
        hdr_row = 5
        for col, h in enumerate(HDR, 1):
            cell             = ws.cell(row=hdr_row, column=col, value=h)
            cell.font        = hdr_font
            cell.fill        = hdr_fill
            cell.alignment   = center
            cell.border      = border
        ws.row_dimensions[hdr_row].height = 22

        # ── Data rows ─────────────────────────────────────────────
        for ei, row in enumerate(rows):
            r     = hdr_row + 1 + ei
            is_alt = ei % 2 == 0
            fill  = alt_fill if is_alt else PatternFill()

            data = [
                ei + 1,
                row["company_name"],
                row["product_name"],
                row["invoice_number"],
                row["invoice_date"],
                row["total_amount"],
                row["paid_date"],
                row["payment_mode"],
                row["amount"],
                row["expense_status"],
                row["total_paid"],
                row["balance"],
            ]

            for col, val in enumerate(data, 1):
                cell           = ws.cell(row=r, column=col, value=val)
                cell.font      = body_font
                cell.fill      = fill
                cell.border    = border
                cell.alignment = center if col in (1, 6, 8, 9, 10) else left

                # ✅ Money formatting
                if col in (6, 9, 11, 12):
                    cell.number_format = money_fmt
                    cell.alignment     = right

                # ✅ Balance colour
                if col == 12:
                    cell.font = red_font if (row["balance"] or 0) > 0 else grn_font

            ws.row_dimensions[r].height = 18

        # ── Grand total row ───────────────────────────────────────
        tot_row = hdr_row + 1 + len(rows)
        tot_data = [
            "", "GRAND TOTAL", "", "", f"{len(rows)} records", "",
            "", "", total_item_amt, "", total_paid, total_bal,
        ]
        for col, val in enumerate(tot_data, 1):
            cell           = ws.cell(row=tot_row, column=col, value=val)
            cell.font      = tot_font
            cell.fill      = tot_fill
            cell.border    = border
            cell.alignment = right if col in (9, 11, 12) else center
            if col in (9, 11, 12):
                cell.number_format = money_fmt
        ws.row_dimensions[tot_row].height = 22

        # ── Column widths ─────────────────────────────────────────
        col_widths = [6, 20, 18, 16, 14, 16, 14, 14, 14, 12, 14, 16]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # ── Freeze header ─────────────────────────────────────────
        ws.freeze_panes = f"A{hdr_row + 1}"

        # ── Auto-filter ───────────────────────────────────────────
        ws.auto_filter.ref = f"A{hdr_row}:L{hdr_row}"

        # ── Stream response ───────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return HttpResponse(
            buf,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="arrolite_expense_report.xlsx"'},
        )